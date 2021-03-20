"""
Напишите программу, которая:
- прочитает содержимое всех конфигурационных файлов из каталога (аналогично лабораторной работе 1.5);
- сгенерирует список всех интерфейсов на всех устройствах, с указанием IP-адресов и масок (аналогично лабораторной работе 1.6);
- сгенерирует адресный план — таблицу со столбцами «Сеть», «Маска» (комбинации сети и маски должны быть уникальными);
- выведет адресный план на экран — в виде, пригодном для чтения, формат вывода определите сами.

Дополнительное задание («со звёздочкой»): сгенерируйте адресный план в виде таблицы MS Excel, воспользуйтесь библиотекой openpyxl (см. лабораторную работу 1.2)
"""
import re
import ipaddress
import glob
import openpyxl

ip_list = []
def func(val):
    ip_mask = re.search('(?<=ip address )(([0-9]{1,3}[.]){3}[0-9]{1,3}) (([0-9]{1,3}[.]){3}[0-9]{1,3})', val)
    if ip_mask:
        temp = ipaddress.IPv4Interface(ip_mask.group(0).replace(' ','/'))
        net = str(temp.network.network_address)
        mask = str(temp.netmask)
        res = net + ' ' + mask
        return res
    else: return False

for file in glob.glob('d:\OneDrive\Docs\Python_cource\config_files\*'):
    with open(file) as f:
        for line in f:
            a = func(line)
            if a:
                ip_list.append(a)
"""
пример для 10.12.226.50/26
ip_list[0].netmask
Out[3]: IPv4Address('255.255.255.192')
ip_list[0].network
Out[4]: IPv4Network('10.12.226.0/26')
ip_list[0].with_netmask
Out[5]: '10.12.226.50/255.255.255.192'
ip_list[0].network.network_address
Out[6]: IPv4Address('10.12.226.0')
ip_list[0].network.prefixlen
Out[7]: 26
"""
ip_list_u = list(set(ip_list))
# print(ip_list_u)
print('+-----------------+-----------------+')
print('| Сеть            | Маска           |')
print('+-----------------+-----------------+')
for i in ip_list_u:
    j = i.split(' ')
    print('| %15s | %15s |' % (j[0],j[1]))
    print('+-----------------+-----------------+')
    
print('Всего сетей :',len(ip_list))
print('Уникальных сетей :',len(ip_list_u))

wb = openpyxl.Workbook()
sheet = wb.active
sheet.cell(row = 1, column = 1).value = 'Сеть'
sheet.cell(row = 1, column = 2).value = 'Маска'
i = 2
while i <= len(ip_list_u)+1:
    j = ip_list_u[i - 2].split(' ')
    sheet.cell(row=i, column=1).value = j[0]
    sheet.cell(row=i, column=2).value = j[1]
    i += 1

wb.save('C:\\Users\\root\\PycharmProjects\\p2ne\\DZ1\\dz.xlsx')
